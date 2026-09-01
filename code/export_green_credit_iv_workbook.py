#!/Users/yumanlou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3
"""Export the IV diagnostics to a readable Excel workbook."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
RESULT_DIR = ROOT / "result" / "tables" / "0716_green_credit_iv"
OUTPUT = RESULT_DIR / "Table_0716_GreenCredit_IV_Diagnostics.xlsx"


def add_frame(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        sheet.append(list(row))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F5F73")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 34)
        sheet.column_dimensions[column[0].column_letter].width = width


def main() -> None:
    first = pd.read_csv(RESULT_DIR / "Table_0716_GreenCredit_IV_FirstStage.csv")
    second = pd.read_csv(RESULT_DIR / "Table_0716_GreenCredit_IV_SecondStage.csv")
    candidates = pd.DataFrame([
        ["行业融资shift-share", "已运行", "联合F最高2.93，KP F低于1.74", "弱工具且排除限制薄弱", "不进入主因果结果"],
        ["银行网络shift-share", "银行冲击已补齐", "仍缺2011年省份×银行网点权重", "相对最有希望", "继续补网点数据"],
        ["绿色金融试验区", "现有", "试点直接影响结果且选择非随机", "不满足排除限制", "只能作替代政策设计"],
        ["政策前贷存比×政策后", "现有", "一般信贷直接影响投资与产出", "不满足排除限制", "不用作工具变量"],
        ["政策文本词频", "现有", "政策后变量且可能直接影响结果", "后定变量", "只作机制和案例证据"],
    ], columns=["候选", "数据状态", "诊断", "识别判断", "处理建议"])

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_frame(workbook, "第一阶段", first)
    add_frame(workbook, "2SLS结果", second)
    add_frame(workbook, "候选工具评估", candidates)

    notes = workbook.create_sheet("说明")
    notes.append(["项目", "内容"])
    notes.append(["内生变量", "绿色信贷代理；绿色信贷代理×政策前资源依赖，两项同时工具化"])
    notes.append(["主探索工具", "政策前六行业融资权重×剔除本省后的全国行业融资变化"])
    notes.append(["样本", "31省，2012—2022年；省份和年份固定效应，标准误按省聚类"])
    notes.append(["结论", "行业融资shift-share为弱工具，2SLS不能解释为因果效应"])
    notes.append(["下一步", "补齐2011年省份×银行网点权重，构造银行网络shift-share"])
    for cell in notes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F5F73")
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 90
    for row in notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(OUTPUT)
    check = load_workbook(OUTPUT, read_only=True)
    expected = {"第一阶段", "2SLS结果", "候选工具评估", "说明"}
    if set(check.sheetnames) != expected:
        raise ValueError(f"Unexpected workbook sheets: {check.sheetnames}")
    print(f"saved={OUTPUT}")
    print(f"sheets={check.sheetnames}")


if __name__ == "__main__":
    main()
