#!/Users/yumanlou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3
"""Build exploratory shift-share instruments for the green-credit proxy.

The instrument combines each province's 2007-2011 fixed exposure to six
energy-intensive industries with leave-one-province-out national changes in
the industries' interest-expense shares. It is useful for first-stage and
sensitivity diagnostics, but national industry financing shocks can affect
energy and pollution outcomes directly, so it is not treated as a clean IV.
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "green_credit" / "2005-2022年绿色信贷水平.XLSX"
INPUT_PANEL = ROOT / "data" / "final_data.1.3.4_did_full_resource_v2_credit_greencredit_0716.csv"
OUTPUT_DIR = ROOT / "data" / "green_credit_iv"
OUTPUT_PANEL = ROOT / "data" / "final_data.1.3.4_did_full_resource_v2_credit_greencredit_iv_0716.csv"

PROVINCE_CN_TO_MAIN = {
    "北京": "Beijing", "天津": "Tianjin", "河北": "Hebei", "山西": "Shanxi",
    "内蒙古": "Neimenggu", "辽宁": "Liaoning", "吉林": "Jilin", "黑龙江": "Heilongjiang",
    "上海": "Shanghai", "江苏": "Jiangsu", "浙江": "Zhejiang", "安徽": "Anhui",
    "福建": "Fujian", "江西": "Jiangxi", "山东": "Shandong", "河南": "Henan",
    "湖北": "Hubei", "湖南": "Hunan", "广东": "Guangdong", "广西": "Guangxi",
    "海南": "Hainan", "重庆": "Chongqing", "四川": "Sichuan", "贵州": "Guizhou",
    "云南": "Yunnan", "西藏": "Xizang", "陕西": "Shaanxi", "甘肃": "Gansu",
    "青海": "Qinghai", "宁夏": "Ningxia", "新疆": "Xinjiang",
}
SECTORS = ["chemical", "petroleum", "power_heat", "ferrous", "nonferrous", "nonmetal"]
BASE_YEARS = list(range(2007, 2012))


def normalize_label(value: object) -> str:
    return str(value or "").replace(" ", "").replace("\u3000", "").strip()


def numeric_cell(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    normalized = str(value).replace("．", ".").replace("，", ",").replace(" ", "")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", normalized.replace(",", ""))
    if not match:
        raise ValueError(f"Cannot parse numeric workbook cell: {value!r}")
    return float(match.group())


def source_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def parse_raw_row(values: tuple[object, ...], year: int) -> dict | None:
    label = normalize_label(values[0])
    if label not in PROVINCE_CN_TO_MAIN:
        return None
    numeric = [numeric_cell(value) for value in values[1:8]]
    total = numeric[0]
    if total <= 0:
        raise ValueError(f"Non-positive industrial interest expense: {label} {year}")
    row = {"province": PROVINCE_CN_TO_MAIN[label], "year": year, "total_interest": total}
    row.update(dict(zip(SECTORS, numeric[1:])))
    return row


def extract_sector_panel() -> pd.DataFrame:
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    rows: list[dict] = []

    sheet = workbook["Sheet3"]
    for values in sheet.iter_rows(min_row=2, max_col=9, values_only=True):
        if not isinstance(values[8], (int, float)):
            continue
        year = int(values[8])
        if year == 2017:
            continue
        row = parse_raw_row(values[:8], year)
        if row:
            rows.append(row)

    for year, sheet_name, start_row in [
        (2017, "2017年插值法填充", 3),
        (2020, "2014-2020计算过程", 5),
        (2021, "2021", 5),
        (2022, "2022年", 5),
    ]:
        sheet = workbook[sheet_name]
        for values in sheet.iter_rows(min_row=start_row, max_col=8, values_only=True):
            row = parse_raw_row(values, year)
            if row:
                rows.append(row)

    data = pd.DataFrame(rows).sort_values(["province", "year"]).reset_index(drop=True)
    # The Hubei 2007 ferrous cell is stored as text ("11．44．") and is excluded
    # by the workbook's reported SUM formula. Preserve that reported convention.
    hubei_2007 = (data["province"] == "Hubei") & (data["year"] == 2007)
    if hubei_2007.sum() != 1:
        raise ValueError("Could not locate the documented Hubei 2007 source anomaly")
    data.loc[hubei_2007, "ferrous"] = 0.0
    expected = 31 * 16
    if len(data) != expected or data["province"].nunique() != 31:
        raise ValueError(f"Expected {expected} rows for 31 provinces, got {len(data)}")
    if set(data["year"]) != set(range(2007, 2023)):
        raise ValueError(f"Unexpected years: {sorted(data['year'].unique())}")
    if data.duplicated(["province", "year"]).any():
        raise ValueError("Duplicate province-year keys in sector data")

    reconstructed = data[SECTORS].sum(axis=1) / data["total_interest"]
    cleaned = pd.read_csv(OUTPUT_DIR.parent / "green_credit" / "green_credit_province_2005_2022.csv")
    check = data[["province", "year"]].copy()
    check["reconstructed"] = reconstructed
    check = check.merge(
        cleaned[["province", "year", "green_credit_high_energy_interest_share"]],
        on=["province", "year"], how="left", validate="one_to_one",
    )
    max_diff = (check["reconstructed"] - check["green_credit_high_energy_interest_share"]).abs().max()
    if max_diff > 1e-12:
        check["difference"] = (check["reconstructed"] - check["green_credit_high_energy_interest_share"]).abs()
        failures = check.nlargest(8, "difference").to_dict(orient="records")
        raise ValueError(
            f"Sector decomposition does not reconcile, max difference={max_diff}; top={failures}"
        )
    return data


def construct_instruments(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    baseline = data[data["year"].isin(BASE_YEARS)].copy()
    base_local = baseline.groupby("province", as_index=True)[SECTORS + ["total_interest"]].sum()
    weights = base_local[SECTORS].div(base_local["total_interest"], axis=0)
    weights.columns = [f"base_share_{sector}" for sector in SECTORS]

    national_by_year = data.groupby("year", as_index=True)[SECTORS + ["total_interest"]].sum()
    records: list[dict] = []
    for province, province_data in data.groupby("province"):
        province_data = province_data.set_index("year").sort_index()
        loo = national_by_year - province_data[SECTORS + ["total_interest"]]
        loo_base = loo.loc[BASE_YEARS].sum()
        base_total_share = loo_base[SECTORS] / loo_base["total_interest"]
        local_weights = weights.loc[province]

        for year, row in province_data.iterrows():
            loo_sector_share = loo.loc[year, SECTORS] / loo.loc[year, "total_interest"]
            relative_shift = loo_sector_share / base_total_share - 1.0
            level_shift = loo_sector_share - base_total_share
            predicted_relative = sum(
                local_weights[f"base_share_{sector}"] * relative_shift[sector]
                for sector in SECTORS
            )
            predicted_level = sum(
                local_weights[f"base_share_{sector}"] * level_shift[sector]
                for sector in SECTORS
            )
            records.append({
                "province": province,
                "year": int(year),
                "iv_bartik_gc_relative_raw": -predicted_relative,
                "iv_bartik_gc_level_raw": -predicted_level,
                "iv_interpolated_flag": int(year == 2017),
            })

    iv = pd.DataFrame(records).sort_values(["province", "year"]).reset_index(drop=True)
    post = iv["year"].between(2012, 2022)
    for raw, standardized in [
        ("iv_bartik_gc_relative_raw", "iv_bartik_gc_relative_z"),
        ("iv_bartik_gc_level_raw", "iv_bartik_gc_level_z"),
    ]:
        mean = iv.loc[post, raw].mean()
        sd = iv.loc[post, raw].std(ddof=1)
        if not sd > 0:
            raise ValueError(f"Instrument has zero variance: {raw}")
        iv[standardized] = (iv[raw] - mean) / sd

    weight_output = weights.reset_index()
    weight_output["base_high_energy_share"] = weight_output.drop(columns="province").sum(axis=1)
    return iv, weight_output


def merge_panel(iv: pd.DataFrame) -> pd.DataFrame:
    panel = pd.read_csv(INPUT_PANEL)
    merged = panel.merge(iv, on=["province", "year"], how="left", validate="one_to_one")
    for stem in ["relative", "level"]:
        merged[f"iv_bartik_gc_{stem}_z_x_resdep"] = (
            merged[f"iv_bartik_gc_{stem}_z"] * merged["resdep_pre"]
        )
    if len(merged) != 744 or merged.duplicated(["province", "year"]).any():
        raise ValueError("Merged panel structure changed")
    merged.to_csv(OUTPUT_PANEL, index=False, encoding="utf-8-sig")
    return merged


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sector = extract_sector_panel()
    iv, weights = construct_instruments(sector)
    merged = merge_panel(iv)

    sector.to_csv(OUTPUT_DIR / "green_credit_sector_interest_2007_2022.csv", index=False, encoding="utf-8-sig")
    weights.to_csv(OUTPUT_DIR / "green_credit_baseline_industry_weights_2007_2011.csv", index=False, encoding="utf-8-sig")
    iv.to_csv(OUTPUT_DIR / "green_credit_shiftshare_iv_2007_2022.csv", index=False, encoding="utf-8-sig")

    post = merged[merged["year"].between(2012, 2022) & merged["green_credit_proxy"].notna()]
    metadata = {
        "source_file": SOURCE.name,
        "source_sha256": source_hash(SOURCE),
        "sector_years": [2007, 2022],
        "baseline_years": BASE_YEARS,
        "province_n": 31,
        "post_iv_observations": int(len(post)),
        "main_instrument": "iv_bartik_gc_relative_z",
        "alternative_instrument": "iv_bartik_gc_level_z",
        "construction": (
            "Negative weighted sum of leave-one-province-out national changes in six "
            "energy-intensive industries' interest-expense shares; weights are fixed "
            "province exposures from 2007-2011."
        ),
        "identification_warning": (
            "National industry financing shocks may directly affect provincial energy use, "
            "pollution, and output. The exclusion restriction is not established, so this "
            "instrument is exploratory rather than a main causal design."
        ),
        "source_anomaly": (
            "Hubei 2007 ferrous interest expense is stored as text '11．44．' and omitted "
            "by the workbook's reported SUM formula; the sector panel sets it to zero to "
            "reconcile exactly with the paper's green-credit proxy."
        ),
        "correlations_2012_2022": {
            "relative_iv_with_green_credit_proxy": float(post["iv_bartik_gc_relative_z"].corr(post["green_credit_proxy"])),
            "level_iv_with_green_credit_proxy": float(post["iv_bartik_gc_level_z"].corr(post["green_credit_proxy"])),
            "relative_with_level_iv": float(post["iv_bartik_gc_relative_z"].corr(post["iv_bartik_gc_level_z"])),
        },
    }
    (OUTPUT_DIR / "green_credit_shiftshare_iv_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(metadata, ensure_ascii=False, indent=2))
    print(f"merged panel={OUTPUT_PANEL}")


if __name__ == "__main__":
    main()
