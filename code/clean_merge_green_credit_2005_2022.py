#!/Users/yumanlou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3
"""Clean the provincial green-credit proxy workbook and merge it to the panel.

The workbook's reported result is the interest-expense share of six
energy-intensive industries. The positive green-credit proxy used here is
one minus that share. The source workbook is never modified.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DEFAULT = ROOT / "data" / "green_credit" / "2005-2022年绿色信贷水平.XLSX"
PANEL_DEFAULT = ROOT / "data" / "final_data.1.3.4_did_full_resource_v2_credit_0716.csv"
OUTPUT_PANEL_DEFAULT = ROOT / "data" / "final_data.1.3.4_did_full_resource_v2_credit_greencredit_0716.csv"
OUTPUT_DIR = ROOT / "data" / "green_credit"

PROVINCE_CN_TO_MAIN = {
    "北京": "Beijing", "天津": "Tianjin", "河北": "Hebei", "山西": "Shanxi",
    "内蒙古": "Neimenggu", "辽宁": "Liaoning", "吉林": "Jilin", "黑龙江": "Heilongjiang",
    "上海": "Shanghai", "江苏": "Jiangsu", "浙江": "Zhejiang", "安徽": "Anhui",
    "福建": "Fujian", "江西": "Jiangxi", "山东": "Shandong", "河南": "Henan",
    "湖北": "Hubei", "湖南": "Hunan", "广东": "Guangdong", "广西": "Guangxi",
    "海南": "Hainan", "重庆": "Chongqing", "四川": "Sichuan", "贵州": "Guizhou",
    "云南": "Yunnan", "西藏": "Xizang", "陕西": "Shaanxi", "甘肃": "Gansu",
    "青海": "Qinghai", "宁夏": "Ningxia", "新疆": "Xinjiang",
}
EXPECTED_YEARS = set(range(2005, 2023))


def normalize_label(value: object) -> str:
    return str(value or "").replace(" ", "").replace("\u3000", "").strip()


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def workbook_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def extract_result(source: Path) -> tuple[list[dict], list[dict], dict]:
    values_book = load_workbook(source, data_only=True, read_only=True)
    formula_book = load_workbook(source, data_only=False, read_only=True)
    values_sheet = values_book["绿色信贷-计算结果"]
    formula_sheet = formula_book["绿色信贷-计算结果"]

    year_columns = {}
    for column in range(2, values_sheet.max_column + 1):
        value = values_sheet.cell(1, column).value
        if isinstance(value, (int, float)) and int(value) in EXPECTED_YEARS:
            year = int(value)
            if year in year_columns:
                raise ValueError(f"Duplicate year column in source workbook: {year}")
            year_columns[year] = column
    if set(year_columns) != EXPECTED_YEARS:
        raise ValueError(f"Unexpected year coverage: {sorted(year_columns)}")

    province_rows = {}
    national_row = None
    for row in range(2, values_sheet.max_row + 1):
        label = normalize_label(values_sheet.cell(row, 1).value)
        if label == "全国":
            national_row = row
        elif label in PROVINCE_CN_TO_MAIN:
            province_rows[PROVINCE_CN_TO_MAIN[label]] = row
    if set(province_rows) != set(PROVINCE_CN_TO_MAIN.values()):
        missing = sorted(set(PROVINCE_CN_TO_MAIN.values()) - set(province_rows))
        raise ValueError(f"Missing provinces in source workbook: {missing}")
    if national_row is None:
        raise ValueError("National row not found")

    province_output = []
    national_output = []
    for province, row in sorted(province_rows.items()):
        for year in sorted(EXPECTED_YEARS):
            column = year_columns[year]
            share = values_sheet.cell(row, column).value
            if not isinstance(share, (int, float)):
                raise ValueError(f"Missing/non-numeric result: {province} {year} {share!r}")
            share = float(share)
            if not 0 <= share <= 1:
                raise ValueError(f"Share outside [0,1]: {province} {year} {share}")
            formula = formula_sheet.cell(row, column).value
            if year == 2017 and "2017年插值法填充" not in str(formula):
                raise ValueError(f"2017 interpolation link missing: {province} {formula!r}")
            province_output.append({
                "province": province,
                "year": year,
                "green_credit_high_energy_interest_share": share,
                "green_credit_proxy": 1.0 - share,
                "green_credit_proxy_pct": 100.0 * (1.0 - share),
                "green_credit_interpolated_flag": int(year == 2017),
            })

    for year in sorted(EXPECTED_YEARS):
        share = values_sheet.cell(national_row, year_columns[year]).value
        if not isinstance(share, (int, float)):
            raise ValueError(f"Missing national result: {year}")
        national_output.append({
            "year": year,
            "green_credit_high_energy_interest_share_national": float(share),
            "green_credit_proxy_national": 1.0 - float(share),
        })

    if len(province_output) != 31 * 18:
        raise ValueError(f"Expected 558 province-year rows, got {len(province_output)}")
    if len({(row["province"], row["year"]) for row in province_output}) != len(province_output):
        raise ValueError("Duplicate province-year key in cleaned output")

    # Reconcile 2021-2022 result cells with raw interest-expense sheets.
    reconciliation = {}
    for year, sheet_name in [(2021, "2021"), (2022, "2022年")]:
        raw_sheet = values_book[sheet_name]
        raw_by_province = {}
        for row in range(5, raw_sheet.max_row + 1):
            label = normalize_label(raw_sheet.cell(row, 1).value)
            if label not in PROVINCE_CN_TO_MAIN:
                continue
            total = raw_sheet.cell(row, 2).value
            sectors = [raw_sheet.cell(row, column).value or 0 for column in range(3, 9)]
            computed = sum(float(value) for value in sectors) / float(total)
            raw_by_province[PROVINCE_CN_TO_MAIN[label]] = computed
        result_by_province = {
            row["province"]: row["green_credit_high_energy_interest_share"]
            for row in province_output if row["year"] == year
        }
        differences = {
            province: abs(result_by_province[province] - raw_by_province[province])
            for province in result_by_province
        }
        reconciliation[str(year)] = {
            "province_n": len(raw_by_province),
            "maximum_absolute_difference": max(differences.values()),
        }
        if len(raw_by_province) != 31 or max(differences.values()) > 1e-12:
            raise ValueError(f"Raw-sheet reconciliation failed for {year}: {reconciliation[str(year)]}")

    metadata = {
        "source_file": source.name,
        "source_sha256": workbook_sha256(source),
        "source_sheet": "绿色信贷-计算结果",
        "source_measure": "六大高耗能行业利息费用/规模以上工业企业利息费用",
        "positive_proxy_formula": "1 - green_credit_high_energy_interest_share",
        "years": [2005, 2022],
        "province_n": 31,
        "province_year_n": len(province_output),
        "interpolated_year": 2017,
        "reconciliation": reconciliation,
        "notes": [
            "The source workbook calls the high-energy interest share a green-credit result.",
            "Higher green_credit_proxy means a lower financing share for six energy-intensive industries.",
            "This is an indirect industrial-interest proxy, not disclosed provincial green-loan balance.",
        ],
    }
    return province_output, national_output, metadata


def merge_panel(panel_path: Path, output_path: Path, green_rows: list[dict]) -> None:
    green = {(row["province"], row["year"]): row for row in green_rows}
    with panel_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        source_fields = list(reader.fieldnames or [])
        panel_rows = list(reader)

    additions = [
        "green_credit_high_energy_interest_share", "green_credit_proxy",
        "green_credit_proxy_pct", "green_credit_interpolated_flag",
        "green_credit_proxy_x_resdep_pre", "green_credit_proxy_x_coalexp_pre",
    ]
    if any(field in source_fields for field in additions):
        raise ValueError("Green-credit variables already exist in input panel")

    seen = set()
    matched = 0
    for row in panel_rows:
        key = (row["province"], int(float(row["year"])))
        if key in seen:
            raise ValueError(f"Duplicate input panel key: {key}")
        seen.add(key)
        source = green.get(key)
        if source is None:
            for field in additions:
                row[field] = ""
            continue
        matched += 1
        for field in additions[:4]:
            row[field] = source[field]
        proxy = float(source["green_credit_proxy"])
        row["green_credit_proxy_x_resdep_pre"] = (
            proxy * float(row["resdep_pre"]) if row.get("resdep_pre", "") != "" else ""
        )
        row["green_credit_proxy_x_coalexp_pre"] = (
            proxy * float(row["coalexp_pre"]) if row.get("coalexp_pre", "") != "" else ""
        )

    if len(panel_rows) != 744 or len(seen) != 744:
        raise ValueError(f"Unexpected panel structure: rows={len(panel_rows)} keys={len(seen)}")
    if matched != 558:
        raise ValueError(f"Expected 558 matched province-year rows, got {matched}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=source_fields + additions)
        writer.writeheader()
        writer.writerows(panel_rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=SOURCE_DEFAULT)
    parser.add_argument("--panel", type=Path, default=PANEL_DEFAULT)
    parser.add_argument("--output-panel", type=Path, default=OUTPUT_PANEL_DEFAULT)
    args = parser.parse_args()

    province_rows, national_rows, metadata = extract_result(args.source)
    write_csv(
        OUTPUT_DIR / "green_credit_province_2005_2022.csv", province_rows,
        ["province", "year", "green_credit_high_energy_interest_share",
         "green_credit_proxy", "green_credit_proxy_pct", "green_credit_interpolated_flag"],
    )
    write_csv(
        OUTPUT_DIR / "green_credit_national_2005_2022.csv", national_rows,
        ["year", "green_credit_high_energy_interest_share_national",
         "green_credit_proxy_national"],
    )
    with (OUTPUT_DIR / "green_credit_metadata.json").open("w", encoding="utf-8") as handle:
        json.dump(metadata, handle, ensure_ascii=False, indent=2)
    merge_panel(args.panel, args.output_panel, province_rows)

    proxies = [row["green_credit_proxy"] for row in province_rows]
    print(f"clean rows={len(province_rows)} years=2005-2022 provinces=31")
    print(f"green_credit_proxy min={min(proxies):.6f} max={max(proxies):.6f}")
    print(f"merged panel={args.output_panel}")
    print(f"sha256={metadata['source_sha256']}")


if __name__ == "__main__":
    main()
